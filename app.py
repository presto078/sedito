import streamlit as st
import pandas as pd
import io
from openpyxl.styles import Font

st.set_page_config(page_title="SedíTo! - Profesionální párování tržeb", page_icon="💳", layout="centered")

st.title("💳 SedíTo! – Kontrola a párování tržeb")
st.write("Profesionální odsouhlasení pokladních dat Datona vůči bankovním terminálům a Amexu s finančním auditem.")

# 1. Nahrání souborů
st.subheader("1. Krok: Nahrání podkladů")

file_pokpol = st.file_uploader("Soubor prodejů Datona (pokpol.csv / xlsx)", type=["csv", "xlsx"])
file_karty = st.file_uploader("Soubor transakcí - KARTY (bankovní výpis)", type=["csv", "xlsx"])
file_amex = st.file_uploader("Soubor transakcí - AMEX", type=["csv", "xlsx"])

def load_df(uploaded_file, skip_rows=0):
    if uploaded_file.name.lower().endswith('.csv'):
        try:
            return pd.read_csv(uploaded_file, skiprows=skip_rows)
        except:
            return pd.read_csv(uploaded_file, skiprows=skip_rows, encoding='cp1250', sep=None, engine='python')
    else:
        return pd.read_excel(uploaded_file, skiprows=skip_rows)

if st.button("🚀 Spustit hloubkovou analýzu", use_container_width=True):
    if not file_pokpol or not file_karty or not file_amex:
        st.error("Prosím, nahrajte všechny 3 požadované soubory.")
    else:
        with st.spinner("Provádím finanční audit tržeb..."):
            try:
                # Načtení dat
                pokpol = load_df(file_pokpol, skip_rows=0)
                karty = load_df(file_karty, skip_rows=5)
                amex = load_df(file_amex, skip_rows=5)
                
                # Převod částek a datumů
                pokpol['Cena'] = pd.to_numeric(pokpol['Cena'], errors='coerce')
                karty['Částka brutto'] = pd.to_numeric(karty['Částka brutto'], errors='coerce')
                amex['Částka brutto'] = pd.to_numeric(amex['Částka brutto'], errors='coerce')
                
                pokpol['dt'] = pd.to_datetime(pokpol['Datum a Čas'], errors='coerce', dayfirst=True)
                karty['dt'] = pd.to_datetime(karty['Čas transakce'], errors='coerce', dayfirst=True)
                amex['dt'] = pd.to_datetime(amex['Čas transakce'], errors='coerce', dayfirst=True)
                
                # Sjednocení bankovních terminálů
                karty['Zdroj'] = 'Karty'
                amex['Zdroj'] = 'Amex'
                term_cols = ['Částka brutto', 'Čas transakce', 'dt', 'Zdroj', 'Číslo karty', 'Typ karty']
                for col in term_cols:
                    if col not in karty.columns: karty[col] = ''
                    if col not in amex.columns: amex[col] = ''
                terminal_all = pd.concat([karty[term_cols], amex[term_cols]], ignore_index=True).sort_values('dt')
                terminal_all['matched'] = False
                
                # Rozdělení pokladny podle sloupce ZpPlat
                pokpol_karty = pokpol[pokpol['ZpPlat'] == 'K'].copy()
                pokpol_ostatni = pokpol[pokpol['ZpPlat'] != 'K'].copy()
                
                # --- 1. KROK: VYRUŠENÍ VNITŘNÍCH STOREN В KASE ---
                pokpol_karty['vnitrni_storno'] = False
                pokpol_k_pos = pokpol_karty[pokpol_karty['Cena'] > 0].copy().sort_values('dt')
                pokpol_k_neg = pokpol_karty[pokpol_karty['Cena'] < 0].copy().sort_values('dt')
                
                storna_rows = []
                for n_idx, n_row in pokpol_k_neg.iterrows():
                    target_amt = abs(n_row['Cena'])
                    candidates = pokpol_k_pos[(pokpol_k_pos['Cena'] == target_amt) & (~pokpol_k_pos['vnitrni_storno'])]
                    if not candidates.empty:
                        time_diffs = (candidates['dt'] - n_row['dt']).abs()
                        time_diffs_12h = ((candidates['dt'] - n_row['dt']).abs() - pd.Timedelta(hours=12)).abs()
                        final_diffs = pd.concat([time_diffs, time_diffs_12h], axis=1).min(axis=1)
                        best_p_idx = final_diffs.idxmin()
                        
                        pokpol_k_pos.loc[best_p_idx, 'vnitrni_storno'] = True
                        storna_rows.append({
                            'Datum Pokladna': n_row['Datum a Čas'],
                            'Doklad CZAK (Storno)': n_row['CZAK'],
                            'Částka Storna': n_row['Cena'],
                            'Původní Doklad CZAK': pokpol_k_pos.loc[best_p_idx, 'CZAK'],
                            'Stav': 'Interně stornováno (V pořádku)'
                        })
                    else:
                        storna_rows.append({
                            'Datum Pokladna': n_row['Datum a Čas'],
                            'Doklad CZAK (Storno)': n_row['CZAK'],
                            'Částka Storna': n_row['Cena'],
                            'Původní Doklad CZAK': 'Nenalezen',
                            'Stav': 'Sirotčí storno (Chybí prodej)'
                        })
                
                pokpol_active_karty = pokpol_k_pos[~pokpol_k_pos['vnitrni_storno']].copy()
                
                # --- 2. KROK: INTELIGENTNÍ PÁROVÁNÍ 1:1 S OPRAVOU AM/PM ČASU ---
                matched_list = []
                unmatched_pokpol = []
                
                for idx, row in pokpol_active_karty.iterrows():
                    amt = row['Cena']
                    candidates = terminal_all[(terminal_all['Částka brutto'] == amt) & (~terminal_all['matched'])]
                    
                    if not candidates.empty:
                        diff_normal = (candidates['dt'] - row['dt']).abs()
                        diff_12h = ((candidates['dt'] - (row['dt'] + pd.Timedelta(hours=12))).abs())
                        combined_diffs = pd.concat([diff_normal, diff_12h], axis=1).min(axis=1)
                        best_idx = combined_diffs.idxmin()
                        
                        terminal_all.loc[best_idx, 'matched'] = True
                        matched_list.append({
                            'Pokladna Datum': row['Datum a Čas'],
                            'Doklad CZAK': row['CZAK'],
                            'Částka Pokladna': amt,
                            'Terminál Čas': terminal_all.loc[best_idx, 'Čas transakce'],
                            'Terminál Částka': terminal_all.loc[best_idx, 'Částka brutto'],
                            'Zdroj': terminal_all.loc[best_idx, 'Zdroj'],
                            'Stav': 'Spárováno'
                        })
                    else:
                        unmatched_pokpol.append(row)
                
                # --- 3. KROK: AUDIT ROZDÍLŮ A FINANČNÍ DOPADY ---
                df_prebyva_terminal = terminal_all[~terminal_all['matched']].copy()
                df_chyby_preklepy_zaměny = []
                
                suma_chybi = 0
                suma_prebyva = 0
                
                # Projdeme zbylé nespárované z kasy a zkusíme u nich napřed najít překlep na terminálu (rozdíl do 200 Kč)
                pokladna_zpracovano = set()
                terminal_zpracovano = set()
                
                for idx, row in unmatched_pokpol:
                    amt_kasa = row['Cena']
                    day_kasa = pd.to_datetime(row['Datum a Čas'], dayfirst=True).date()
                    
                    for t_idx, t_row in df_prebyva_terminal.iterrows():
                        if t_idx in terminal_zpracovano: continue
                        amt_banka = t_row['Částka brutto']
                        day_banka = t_row['dt'].date()
                        
                        # Pokud je to stejný den a rozdíl je drobný (např. do 200 Kč) -> je to PŘEKLEP OBSLUHY
                        if day_kasa == day_banka and abs(amt_kasa - amt_banka) <= 200:
                            rozdil = amt_banka - amt_kasa
                            df_chyby_preklepy_zaměny.append({
                                'Typ neshody': '✍️ Překlep obsluhy na terminálu',
                                'Datum / Čas': row['Datum a Čas'],
                                'Částka v Kase': amt_kasa,
                                'Částka v Bance': amt_banka,
                                'Finanční Dopad': rozdil,
                                'Doklad / Karta': row['CZAK'],
                                'Poznámka': f"V kase zapsáno {amt_kasa} Kč, ale na terminálu strženo {amt_banka} Kč. Účetní musí RUČNĚ upravit typ/částku v hotovosti o {rozdil} Kč!"
                            })
                            pokladna_zpracovano.add(row['CZAK'])
                            terminal_zpracovano.add(t_idx)
                            terminal_all.loc[t_idx, 'matched'] = True
                            if rozdil < 0: suma_chybi += abs(rozdil)
                            else: suma_prebyva += rozdil
                            break
                
                # Ty z kasy, které nebyly překlepem, jsou čisté chybějící platby (ztráty)
                for row in unmatched_pokpol:
                    if row['CZAK'] in pokladna_zpracovano: continue
                    df_chyby_preklepy_zaměny.append({
                        'Typ neshody': '❌ Chybí na terminálu (Ztráta)',
                        'Datum / Čas': row['Datum a Čas'],
                        'Částka v Kase': row['Cena'],
                        'Částka v Bance': 0,
                        'Finanční Dopad': -row['Cena'],
                        'Doklad / Karta': row['CZAK'],
                        'Poznámka': 'Zavřeno na kartu, ale zákazník neodpípl / transakce neprošla vůbec.'
                    })
                    suma_chybi += row['Cena']
                
                # Ty z banky, které nebyly překlepem ani záměnou s hotovostí
                for t_idx, t_row in df_prebyva_terminal.iterrows():
                    if t_idx in terminal_zpracovano: continue
                    amt = t_row['Částka brutto']
                    t_day = t_row['dt'].date()
                    
                    # Kontrola záměny za hotovost
                    pokpol_ostatni['date'] = pokpol_ostatni['dt'].dt.date
                    hot_cand = pokpol_ostatni[(pokpol_ostatni['Cena'] == amt) & (pokpol_ostatni['date'] == t_day)]
                    
                    if not hot_cand.empty:
                        best_h_row = hot_cand.iloc[0]
                        df_chyby_preklepy_zaměny.append({
                            'Typ neshody': '⚠️ Záměna: Karta místo HOTOVOSTI',
                            'Datum / Čas': best_h_row['Datum a Čas'],
                            'Částka v Kase': best_h_row['Cena'],
                            'Částka v Bance': amt,
                            'Finanční Dopad': 0,
                            'Doklad / Karta': best_h_row['CZAK'],
                            'Poznámka': f"V kase zapsáno jako Hotovost, ale na terminálu prošla karta. Účetní převede částku {amt} Kč z hotovosti na karty."
                        })
                        terminal_all.loc[t_idx, 'matched'] = True
                        continue
                        
                    # Úplně cizí přebytek v bance
                    df_chyby_preklepy_zaměny.append({
                        'Typ neshody': '💰 Přebývá na terminálu (Nezadáno)',
                        'Datum / Čas': t_row['Čas transakce'],
                        'Částka v Kase': 0,
                        'Částka v Bance': amt,
                        'Finanční Dopad': amt,
                        'Doklad / Karta': t_row['Číslo karty'],
                        'Poznámka': f"Peníze jsou v bance ({t_row['Zdroj']}), ale v kase chybí jakýkoliv doklad."
                    })
                    suma_prebyva += amt
                
                df_neshody_final = pd.DataFrame(df_chyby_preklepy_zaměny)
                df_matched = pd.DataFrame(matched_list)
                df_storna_final = pd.DataFrame(storna_rows)
                
                cisty_rozdil = int(suma_chybi - suma_prebyva)
                
                st.success("Hloubkový audit tržeb dokončen!")
                
                col1, col2 = st.columns(2)
                col1.metric("Celkový rozdíl (Účetní saldo)", f"{cisty_rozdil} Kč")
                col2.metric("Úspěšně spárované tržby", f"{len(df_matched)} ks")
                
                # --- EXPORT DO EXCELU S FINANČNÍM SHRNUTÍM DOLE ---
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_neshody_final.to_excel(writer, sheet_name='HLAVNÍ ROZDÍLY A CHYBY', index=False)
                    
                    worksheet = writer.sheets['HLAVNÍ ROZDÍLY A CHYBY']
                    start_row = len(df_neshody_final) + 3
                    bold_font = Font(bold=True)
                    
                    worksheet.cell(row=start_row, column=1, value="📊 FINANČNÍ SHRNUTÍ AUDITU (PRO ÚČETNÍ)").font = bold_font
                    worksheet.cell(row=start_row+1, column=1, value="Celkem chybí v bance (Manko v kase):")
                    worksheet.cell(row=start_row+1, column=3, value=float(suma_chybi))
                    
                    worksheet.cell(row=start_row+2, column=1, value="Celkem přebývá v bance (Nezadáno v kase):")
                    worksheet.cell(row=start_row+2, column=3, value=float(suma_prebyva))
                    
                    worksheet.cell(row=start_row+3, column=1, value="VÝSLEDNÉ ÚČETNÍ SALDO (ROZDÍL):").font = bold_font
                    worksheet.cell(row=start_row+3, column=3, value=float(cisty_rozdil)).font = bold_font
                    
                    if not df_matched.empty:
                        df_matched.to_excel(writer, sheet_name='V pořádku spárované (1-1)', index=False)
                    if not df_storna_final.empty:
                        df_storna_final.to_excel(writer, sheet_name='Vnitřní storna v kase', index=False)
                        
                excel_data = output.getvalue()
                st.subheader("2. Krok: Stažení kompletního auditu")
                st.download_button(
                    label="📥 Stáhnout pročištěný Excel s přehledem salda",
                    data=excel_data,
                    file_name="Kompletni_Audit_Trzeb_Datona.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Chyba při hloubkové analýze: {str(e)}")
